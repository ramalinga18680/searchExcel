# A simple program to read a column from an excel file and use the data to search the content in another excel file
# and on finding the content in the excel file get the additional information from other columns and write to a new sheet
# TODO : convert this into a class and make it re-usable
import sys

sys.path.append('/usr/lib/python3/dist-packages')
import openpyxl


class ExcelFileSearcher:
    def __repr__(self):
        self.source_file = ""
        self.file_to_search = ""
        self.sheet_to_search = None
        self.sheet_to_write = None

    def get_search_item(self, key, value):
        for row in range(1, self.sheet_to_search.max_row + 1):
            for column in key:
                cell_name = "{}{}".format(column, row)
                cell_name2 = "{}{}".format(value, row)
                print("cell position {} has value {} second row value is {}".format(cell_name, self.sheet_to_search[
                    cell_name].value, self.sheet_to_search[cell_name2].value))

    def search_sheet(self, lookingfor):
        # print("searching for", lookingfor)
        for row in range(1, self.sheet_to_search.max_row + 1):
            for column in range(1, self.sheet_to_search.max_column + 1):
                cell_name = "{}{}".format(column, row)
                if lookingfor == file_reader.sheet_to_search.cell(row, column).value:
                    print("cell found  at row{} and column {}".format(row, column))
                    return row, column
        return None, None

    def search_row(self, lookingfor, rownum):
        # print("searching for", lookingfor)
        for column in range(1, self.sheet_to_search.max_column + 1):
            cell_name = "{}{}".format(column, rownum)
            if lookingfor == self.sheet_to_search.cell(rownum, column).value:
                print("cell found  at row{} and column {}".format(rownum, column))
                return rownum, column
        return None, None

    def search_column(self, looking_for, col_num):
        # print("searching for", looking_for)
        for row in range(1, self.sheet_to_search.max_row + 1):
            cell_name = "{}{}".format(col_num, row)
            if looking_for == self.sheet_to_search.cell(row, col_num).value:
                print("cell found  at row{} and column {}".format(row, col_num))
                return row, col_num
        return None, None

    def fill_column(self, col_num, fill_value):

        for row in range(1, self.sheet_to_search.max_row + 1):
            self.sheet_to_write.cell(row, col_num).value = fill_value

    def remove_space_newline(self, cell_value):
        value = cell_value.strip().replace(' ', '').replace("\n", "")
        return value

    def search_value(self, cell_value, search_in, col_to_lookup):
        towrite = []
        row, column = self.search_column(cell_value, search_in)
        towrite.append(cell_value)
        if row is not None and column is not None:
            cells_to_print = []
            for column_to_lookup in col_to_lookup:
                cells_to_print.append(file_reader.sheet_to_search.cell(row, column_to_lookup))

            for k in range(0, len(cells_to_print)):
                towrite.append(cells_to_print[k].value);

        else:
            for column_to_lookup in col_to_lookup:
                towrite.append("NA");
        return towrite

    def create_write_value(self, to_write_list, to_write_len):
        row_value_to_write = [''] * to_write_len
        for to_write in to_write_list:
            for val in range(0, to_write_len):
                row_value_to_write[val] = row_value_to_write[val] + to_write[val] + "\n"

        self.writerow(row_value_to_write, 2)

    def process_cell_value(self, cell_value, search_in, col_to_lookup):
        to_write_list = []
        for cell in cell_value:
            cell_val = self.remove_space_newline(cell)
            # print(cell_val)
            to_write = self.search_value(cell_val, search_in, col_to_lookup)
            to_write_len = len(to_write)
            to_write_list.append(to_write)
        self.create_write_value(to_write_list, to_write_len)

        # for val in to_write:
        # print(val)

    def on_read_source_cell_data(self, cell_value, search_in, col_to_lookup):
        towrite = []

        # cell_value = self.remove_space_newline(cell_value)
        # print(cell_value)
        cell_value = cell_value.split(',')

        # print(len(cell_value))

        self.process_cell_value(cell_value, search_in, col_to_lookup)

    def load_source_sheet(self, func, searchin, col_to_lookup):
        # eti_list = []
        src_active_sheet = self.source_file[sys.argv[8]]
        col_lookup = openpyxl.utils.column_index_from_string(sys.argv[6])
        # start from range 2 to avoid taking the header row
        for row in range(2, src_active_sheet.max_row + 1):
            value_in_cell = src_active_sheet.cell(row, col_lookup).value
            if value_in_cell is not None:
                # cell_value=self.remove_space_newline(value_in_cell)
                func(value_in_cell, searchin, col_to_lookup)
                # eti_list.append(value_in_cell)
        # return eti_list

    def writerow(self, data, row_to_write):
        '''
        i = 0
        max_range: int = 1 + len(data)
        for col in range(1, max_range):
            self.sheet_to_write.cell(row_to_write, col).value = data[i]
            i += 1
            self.source_file.save(sys.argv[7])
        '''
        self.sheet_to_write.append(data)
        self.source_file.save(sys.argv[7])


# Press the green button in the gutter to run the script.
# argv[1] - excel sheet path and name
# argv[2] - sheet to look in the excel sheet
# argv[3] - item to search for in  the sheet
# argv[4] - search the item in the column
# argv[5] - row to print when the item searched item is found
# argv[6] - column to search in the source sheet
# argv[7] - source excel sheet to pick the ETI ID's
# argv[8] Sheet to look from the source workbook
# usage python3 main.py /home/venkata/PycharmProjects/SearchExcel/testdump.xlsx Delivered "ETI-T100453" M H N
# python3 main.py /home/venkata/PycharmProjects/SearchExcel/testdump.xlsx Delivered "ETI-T100453" M H N  /home/venkata/PycharmProjects/SearchExcel/test.xlsx Automate
# python3 main.py /home/venkata/PycharmProjects/SearchExcel/testdump.xlsx Delivered "ETI-T100453" M [H,N] 2  /home/venkata/PycharmProjects/SearchExcel/test.xlsx Automate
# python3 main.py /home/venkata/PycharmProjects/SearchExcel/testdump.xlsx Delivered "ABCDEF" M [H,N] I  /home/venkata/PycharmProjects/SearchExcel/test.xlsx Automate
#
def extract_list_argument():
    n = len(sys.argv[5])
    a = sys.argv[5][1:n - 1]
    a = a.split(',')
    return a


if __name__ == '__main__':
    print("File to open is {}".format(sys.argv[1]))
    print("Sheet  to open is {}".format(sys.argv[2]))
    print("Item to look for {}".format(sys.argv[3]))
    write_row = int(2)

    # open the file from where the search has to be done
    file_reader = ExcelFileSearcher()
    file_reader.file_to_search = openpyxl.load_workbook(sys.argv[1])
    file_reader.sheet_to_search = file_reader.file_to_search[sys.argv[2]]

    print("Current sheet name is {} ,max rows is {} max columns is {}".format(file_reader.sheet_to_search,
                                                                              file_reader.sheet_to_search.max_row,
                                                                              file_reader.sheet_to_search.max_column))
    # open the file from where the input are to be taken
    file_reader.source_file = openpyxl.load_workbook(sys.argv[7])
    # create a sheet at index 0 to write the search results
    file_reader.sheet_to_write = file_reader.source_file.create_sheet(title='Analysis', index=0)

    # print("src_list length is ", len(src_list))

    l_search_in = openpyxl.utils.column_index_from_string(sys.argv[4])
    col_list = extract_list_argument()
    '''
     for col_alpha in col_list:
        print('col_alpha is ',col_alpha)
    '''
    local_col_to_lookup = []
    for col_alpha in col_list:
        local_col_to_lookup.append(openpyxl.utils.column_index_from_string(col_alpha))

    # parse the file and get the data to be searched as a list
    file_reader.load_source_sheet(file_reader.on_read_source_cell_data, l_search_in, local_col_to_lookup)

    file_reader.file_to_search.close()
    file_reader.source_file.close()

# See PyCharm help at https://www.jetbrains.com/help/pycharm/
