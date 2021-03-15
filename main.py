# A simple program to read a column from an excel file and use the data to search the content in another excel file
# and on finding the content in the excel file get the additional information from other columns and write to a new sheet
# TODO : convert this into a class and make it re-usable
import sys

sys.path.append('/usr/lib/python3/dist-packages')
import openpyxl


class ExcelFileSearcher:
    def __repr__(self):
        self.source_file = ""
        self.file_to_search = ""
        self.sheet_to_search = None
        self.sheet_to_write = None

    def get_search_item(self, key, value):
        for row in range(1, self.sheet_to_search.max_row + 1):
            for column in key:
                cell_name = "{}{}".format(column, row)
                cell_name2 = "{}{}".format(value, row)
                print("cell position {} has value {} second row value is {}".format(cell_name, self.sheet_to_search[
                    cell_name].value, self.sheet_to_search[cell_name2].value))

    def search_sheet(self, lookingfor):
        print("searching for", lookingfor)
        for row in range(1, self.sheet_to_search.max_row + 1):
            for column in range(1, self.sheet_to_search.max_column + 1):
                cell_name = "{}{}".format(column, row)
                if lookingfor == file_reader.sheet_to_search.cell(row, column).value:
                    print("cell found  at row{} and column {}".format(row, column))
                    return row, column
        return None, None

    def search_row(self, lookingfor, rownum):
        print("searching for", lookingfor)
        for column in range(1, self.sheet_to_search.max_column + 1):
            cell_name = "{}{}".format(column, rownum)
            if lookingfor == self.sheet_to_search.cell(rownum, column).value:
                print("cell found  at row{} and column {}".format(rownum, column))
                return rownum, column
        return None, None

    def search_column(self, looking_for, col_num):
        print("searching for", looking_for)
        for row in range(1, self.sheet_to_search.max_row + 1):
            cell_name = "{}{}".format(col_num, row)
            if looking_for == self.sheet_to_search.cell(row, col_num).value:
                print("cell found  at row{} and column {}".format(row, col_num))
                return row, col_num
        return None, None

    def load_source_sheet(self):
        eti_list = []
        src_active_sheet = self.source_file[sys.argv[8]]
        col_lookup = openpyxl.utils.column_index_from_string('I')
        # start from range 2 to avoid taking the header row
        for row in range(2, src_active_sheet.max_row + 1):
            value_in_cell = src_active_sheet.cell(row, col_lookup).value
            if value_in_cell is not None:
                eti_list.append(value_in_cell)
        return eti_list

    def writerow(self, data, row_to_write):
        i = 0
        max_range: int = 2 + len(data)
        for col in range(2, max_range):
            self.sheet_to_write.cell(row_to_write, col).value = data[i]
            i += 1
            self.source_file.save(sys.argv[7])


# Press the green button in the gutter to run the script.
# argv[1] - excel sheet path and name
# argv[2] - sheet to look in the excel sheet
# argv[3] - item to search for in  the sheet
# argv[4] - search the item in the column
# argv[5],argv[6] row to print when the item searched item is found
# argv[7] - source excel sheet to pick the ETI ID's
# argv[8] Sheet to look from the source workbook
# usage python3 main.py /home/venkata/PycharmProjects/SearchExcel/testdump.xlsx Delivered "ETI-T100453" M H N
# python3 main.py /home/venkata/PycharmProjects/SearchExcel/testdump.xlsx Delivered "ETI-T100453" M H N  /home/venkata/PycharmProjects/SearchExcel/test.xlsx Automate


if __name__ == '__main__':
    print("File to open is {}".format(sys.argv[1]))
    print("Sheet  to open is {}".format(sys.argv[2]))
    print("Item to look for {}".format(sys.argv[3]))

    write_row = int(2)
    # open the file from where the search has to be done
    file_reader = ExcelFileSearcher()
    file_reader.file_to_search = openpyxl.load_workbook(sys.argv[1])
    file_reader.sheet_to_search = file_reader.file_to_search[sys.argv[2]]

    print("Current sheet name is {} ,max rows is {} max columns is {}".format(file_reader.sheet_to_search,
                                                                              file_reader.sheet_to_search.max_row,
                                                                              file_reader.sheet_to_search.max_column))
    # open the file from where the input are to be taken
    file_reader.source_file = openpyxl.load_workbook(sys.argv[7])
    # create a sheet at index 0 to write the search results
    file_reader.sheet_to_write = file_reader.source_file.create_sheet(title='Analysis', index=0)
    # parse the file and get the data to be searched as a list
    src_list = file_reader.load_source_sheet()
    print("srclist length is ", len(src_list))

    searchin = openpyxl.utils.column_index_from_string(sys.argv[4])

    # iterate over the srclist and start searching
    for index in src_list:
        # for index in issuelist:
        row, column = file_reader.search_column(index, searchin)
        if row is not None and column is not None:
            col_to_look1 = openpyxl.utils.column_index_from_string(sys.argv[5])
            col_to_look2 = openpyxl.utils.column_index_from_string(sys.argv[6])
            cell_to_print1 = file_reader.sheet_to_search.cell(row, col_to_look1)
            cell_to_print2 = file_reader.sheet_to_search.cell(row, col_to_look2)
            towrite = (index, cell_to_print1.value, cell_to_print2.value)
            file_reader.writerow(towrite, write_row)
            write_row += 1
        else:
            towrite = (index, "TEST_CASE_NOT_FOUND", "CANNOT_DETERMINE_AUTOMATION_STATUS")
            file_reader.writerow(towrite, write_row)
            write_row += 1


# See PyCharm help at https://www.jetbrains.com/help/pycharm/
